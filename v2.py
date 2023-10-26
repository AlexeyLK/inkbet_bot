import openpyxl
import os
import telebot
from telebot import types
import time

bot_token = '6343291110:AAGwqZ6Fg1FGfh32qT3JxgGb3uqvjKPRmS8'  # Вкажіть токен вашого бота
bot = telebot.TeleBot(bot_token)
mark = []
mark.append("some data")

admin_user = '329798696'

users_file = 'users_info.xlsx'
bet_list = 'bet_list.xlsx'
history = 'history.xlsx'
user_data = {}
user_state = {}
# Первинні кнопки
primary_menu = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
btn_help = types.KeyboardButton("Обратная связь")
btn_int = types.KeyboardButton("Инструкция")
btn_menu = types.KeyboardButton("Вызов меню")
btn_info = types.KeyboardButton("Информация о счёте")
primary_menu.add(btn_menu, btn_int, btn_help, btn_info)

# Кнопки для авторизации
markup = types.InlineKeyboardMarkup(row_width=2)
itembtn1 = types.InlineKeyboardButton('Регистрация', callback_data='register')
itembtn2 = types.InlineKeyboardButton('Логин', callback_data='login')
markup.add(itembtn1, itembtn2)

# Кнопки меню
menu_markup = types.InlineKeyboardMarkup(row_width=1)
# m_btn0 = types.InlineKeyboardButton("Посмортеть баланс", callback_data='balance')
m_btn1 = types.InlineKeyboardButton('История ставок', callback_data='history')
# m_btn2 = types.InlineKeyboardButton('Депозит', callback_data='deposit')
m_btn3 = types.InlineKeyboardButton('Сделать ставку', callback_data='view_line')
# m_btn4 = types.InlineKeyboardButton('Вывод', callback_data='out')
m_btn6 = types.InlineKeyboardButton("Мой Баланс", callback_data='balance3')

# menu_markup.add(m_btn0, m_btn1, m_btn2, m_btn3, m_btn4, m_btn6)
menu_markup.add(m_btn6, m_btn3, m_btn1)

# Кнопки выбора дисциплины
menu_sportlist = types.InlineKeyboardMarkup(row_width=2)
spo_btn1 = types.InlineKeyboardButton('Dota 2', callback_data='dota2')
spo_btn2 = types.InlineKeyboardButton('CS 2', callback_data='cs2')
btn_back = types.InlineKeyboardButton("Назад", callback_data="back_s")
menu_sportlist.add(spo_btn1, spo_btn2)
menu_sportlist.add(btn_back)

# Кнопки баланса
menu_balance = types.InlineKeyboardMarkup(row_width=2)
b_btn1 = types.InlineKeyboardButton("Посмортеть баланс", callback_data='balance')
b_btn2 = types.InlineKeyboardButton('Депозит', callback_data='deposit')
b_btn3 = types.InlineKeyboardButton('Вывод', callback_data='out')
b_btn4 = types.InlineKeyboardButton('История переводов ', callback_data='hist_out')
btn_back = types.InlineKeyboardButton("Назад", callback_data="back_s")
menu_balance.add(b_btn1, b_btn2, b_btn3, b_btn4)
menu_balance.add(btn_back)

# Кнопки истории вывода ввода
menu_out = types.InlineKeyboardMarkup(row_width=2)
out_btn1 = types.InlineKeyboardButton("История депозитов", callback_data='dep_hist')
out_btn2 = types.InlineKeyboardButton('История выводов', callback_data='out_hist')
btn_back_to_bal = types.InlineKeyboardButton("Назад", callback_data="back_to_bal")
menu_out.add(out_btn1, out_btn2)
menu_out.add(btn_back_to_bal)

# Кнопки депозита
menu_deposit = types.InlineKeyboardMarkup(row_width=2)
d_btn1 = types.InlineKeyboardButton('USDT TRC20', callback_data='usdt')
d_btn2 = types.InlineKeyboardButton('Перевод на карту', callback_data='carta')
btn_back_dep = types.InlineKeyboardButton("Назад", callback_data="back_dep")
menu_deposit.add(d_btn1, d_btn2)
menu_deposit.add(btn_back_dep)

# Кнопки вывода на карту
menu_card = types.InlineKeyboardMarkup(row_width=2)
car_btn1 = types.InlineKeyboardButton('Сбербанк', callback_data='sber')
car_btn2 = types.InlineKeyboardButton('Тинькофф', callback_data='tink')
btn_back_to_card = types.InlineKeyboardButton("Назад", callback_data="back_to_cards")

menu_card.add(car_btn1, car_btn2)
menu_card.add(btn_back_to_card)

tagret_m = types.InlineKeyboardMarkup(row_width=2)
pay_usdt = types.InlineKeyboardButton("Оплатил", callback_data='oplata')
btn_back_to_card = types.InlineKeyboardButton("Назад", callback_data="back_to_cards")
tagret_m.add(btn_back_to_card, pay_usdt)
# Адмін батон
menu_adm = types.InlineKeyboardMarkup(row_width=2)

# Перевірка наявності файлу користувачів
def check_users_file():
    if not os.path.exists(users_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['user_id', 'balance'])
        wb.save(users_file)
    return
# Перевірка наявності користувача у файлі
def check_user_in_excel(user_id):
    check_users_file()
    wb = openpyxl.load_workbook(users_file)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == user_id:
            return True
    return False

# Реєстрація нового користувача у файлі
def register_user_in_excel(user_id):
    check_users_file()
    if not check_user_in_excel(user_id):
        wb = openpyxl.load_workbook(users_file)
        ws = wb.active
        ws.append([user_id, 50000])
        wb.save(users_file)
    return

def view_line():
    wb = openpyxl.load_workbook(bet_list)
    ws = wb.active
    data_line = {}
    for row in ws.iter_rows(values_only=True):
        key = row[4]
        value = [row[0], row[1], row[2], row[3], row[4]]  # Используйте квадратные скобки для создания списка
        data_line[key] = value

    return data_line

def save_data_to_excel(data, user_id, file_path):
    print(data)
    try:
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['User ID', 'Ставка', 'Сума'])
            wb.save(file_path)
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            data_keys = data.keys()
            key1, key2 = list(data_keys)
            print(key1, key2)
            ws.append([user_id, data[key1], data[key2]])
            wb.save(file_path)
    except Exception as e:
        print(str(e))
def update_balance_xlsx(user_id, amount):
    print(user_id)
    wb = openpyxl.load_workbook(users_file)
    ws = wb.active
    row_number = 2  # Начальное значение номера строки

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] == user_id:
            new_balance = float(row[1]) - float(amount)
            print(new_balance)
            if new_balance < 0:
                return False
            cell = ws.cell(row=row_number, column=2)  # Доступ к ячейке для обновления баланса
            cell.value = new_balance  # Обновление баланса в ячейке

        row_number += 1  # Увеличение номера строки, если пользователь не найден

    wb.save(users_file)  # Сохранение файла
    wb.close()  # Закрытие книги
    return

def check_balance(user_id):
    wb = openpyxl.load_workbook(users_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] == user_id:
            return row[1] 
    return None
mes_info = ""
@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_name = message.from_user.first_name
    user_nickname = message.from_user.username if message.from_user.username else "Нет никнейма"
    iser_id = message.from_user.id
    hi_mes = f"Добро пожаловать, {user_name}!\n"
    hi_mes_ = f"""Желаешь делать ставки  с комфортом и без лишних хлопот? \n
Тогда тебе к нам! Открой новые возможности для успешных ставок  прямо сейчас! 🚀🚀🚀

📌 Наш бот предоставляет уникальный функционал для успешных ставок в мире киберспорта прямо в телеграмм\n
📌 Пополняй баланс любым удобным способом и начинай игру вместе с нами\n
📌 Ставь на понравившейся исход и выигрывай!\n\n
Удачных ставок! 💸💸💸"""
    
    mes_info = f"""🙋🏻‍♂️ Ник: {user_nickname}
👤 ID: {iser_id}
🎰 Активных ставок: в разработке"""
    bot.send_message(message.chat.id, hi_mes + hi_mes_, reply_markup=primary_menu)
    bot.send_message(message.chat.id, "Выберите опцию", reply_markup=markup)
    return




@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    user_id = call.message.chat.id
    if call.data == 'register':
        try:
            register_user_in_excel(user_id)
            if user_id not in user_data:
                user_data[user_id] = {}  # Добавление ключа в user_data
            bot.answer_callback_query(call.id, "Успешная регистрация.")
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Главное меню 🏠", reply_markup=menu_markup)
        except:
            bot.send_message(user_id, 'Ошибка регистрации')
    elif call.data == 'login':
        if check_user_in_excel(user_id):
            bot.answer_callback_query(call.id, "Вы успешно вошли в систему.")
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Главное меню 🏠", reply_markup=menu_markup)
        else:
            bot.answer_callback_query(call.id, "Вы не зарегистрированы.")
# Оброрка депозиту
# Вивід та логіка ставок
    elif call.data == 'balance':
        try:
            balance = check_balance(user_id)
            if balance is not None:
                bot.send_message(user_id, f"🏦 Ваш текущий баланс: {balance} 🏦")
            else:
                bot.send_message(user_id, "Пользователь не зарегистрирован.")
        except:
            bot.send_message(user_id, "Помилка виводу балансу")
    elif call.data == 'view_line':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Выберите киберспортивную дисциплину", reply_markup=menu_sportlist)
    elif call.data == 'dota2':
        try:
            data = view_line()
            # for key in data.keys():
            #     print(len(data[key][4]))
            #     print(data[key][4][0:4])
            view_markup = types.InlineKeyboardMarkup(row_width=1)
            for key in data.keys():
                if len(data[key]) >= 3:  # Проверяем, что есть достаточно данных
                    name1, name2 = data[key][0], data[key][2]
                    if data[key][4]:
                        if key != list(data.keys())[-1] and data[key][4][0:4] == "яяяя":
                            view_markup.add(types.InlineKeyboardButton(f'{name1}   🆚   {name2}', callback_data=f"bet_{key}"))
            view_markup.add(types.InlineKeyboardButton("Назад", callback_data="back_s"))
            
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Ставки', reply_markup=view_markup)
        except:
            bot.send_message(user_id, "Ошибка вывода линии")
            menu_sportlist
    elif call.data == 'cs2':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Ставки по кс2 временно не доступны", reply_markup=menu_markup)
    elif call.data.startswith('bet__'):
        print("")
    elif call.data.startswith('bet_'):
        # try:
        if 1 == 1:
            
            key = call.data[4:]
            mark_ = key[0] + key[4]
            data = view_line()
            menu_b = types.InlineKeyboardMarkup(row_width=2)
            print(1)

            print("here")
            # print(data[mark_][0])
            # print(data[mark_][1])

            # print(data[mark_][2])
            # print(data[mark_][3])

            # print("here------------")
            # print("len" , len(data[key]))
            # print("data[key][4] ", data[key][4])
            for key in data.keys():
                if key:
                    if key[0:2] == mark_ or (key[-1] == mark_[1] and len(key) > 4):
                        print("***********")
                        print(key)
                        print(mark_)
                        print("***********")
                        btn_1 = types.InlineKeyboardButton(f"{data[key][0]} {data[key][1]}", callback_data=f'pay{key}_{data[key][0]}')
                        btn_2 = types.InlineKeyboardButton(f"{data[key][2]} {data[key][3]}", callback_data=f'pay{key}_{data[key][2]}')
                        menu_b.add(btn_1, btn_2)
            
            btn_back = types.InlineKeyboardButton("Назад", callback_data="back")
            menu_b.add(btn_back)


            m_text = "Cделать ставку\n\n 🎮 Выбран вид спорта - Dota 2 \n 🏆 Выбран турнир - TI 2023 \n 🛡 Выберите матч:"
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=m_text, reply_markup=menu_b)
        # except:
        #     bot.send_message(user_id, 'Помилка виводу bet')
    elif call.data.startswith('pay'):
        try:
            s = call.data[3:].split('_')
            data = view_line()
            key = s[0]
            value = s[1]
            name_index = data[key].index(value)  # Получение индекса имени в списке
            
            select_menu = types.InlineKeyboardMarkup(row_width=2)
            btn_pay = types.InlineKeyboardButton('Сделать ставку', callback_data=f"betpay{value} {data[key][name_index + 1]}")
            btn_back = types.InlineKeyboardButton("Назад", callback_data="back")
            select_menu.add(btn_back, btn_pay)

            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f'{value}-{data[key][name_index + 1]}', reply_markup=select_menu)
        except:
            bot.send_message(user_id, 'Помилка виводу pay')
    elif call.data.startswith('betpay'):
        try:
            s = call.data[6:]
            value = {"Ставка": s}
            user_data[user_id] = value

            bot.send_message(user_id, 'Введите сумму')

            @bot.message_handler(func=lambda message: message.text.startswith(''))
            def handle_betpay_message(message):
                try:
                    user_id = message.from_user.id  # Получение уникального идентификатора пользователя
                    user_data[user_id]['updata'] = float(message.text[0:])
                    amount = user_data[user_id]['updata']
                    
                    if update_balance_xlsx(user_id, amount) is False:
                        balance_ = check_balance(user_id)
                        text_= "Недостаточно средств, Ваш баланс: " + str(balance_)
                        bot.send_message(user_id, text_)

                    else:
                        save_data_to_excel(user_data[user_id], user_id, history)

                        print(">>>>", user_data[user_id]['updata'])

                        bot.send_message(user_id, f"Главное меню 🏠\n\n Ставка: {s[:-1]} \n Кеф: {s[-1]} \n Сумма ставки: {amount} \n Желаем удачи!", reply_markup=menu_markup)
                except:
                    bot.send_message(user_id, "Пожалуйства введите правильную сумму ставки")

        except Exception as e:
            bot.send_message(user_id, f'Помилка: {e}')
    elif call.data == 'deposit':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Виберите способ оплаты", reply_markup=menu_deposit)
    elif call.data == 'usdt':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f'Номер кошелька: 1232131 \nВ заметках укажите {user_id} (либо напишите менеджеру @СЮДА), \nНажмите на кнопку Оплатил и Ваш счёт будет пополнен на переведённую сумму в течении 10-15 минут', reply_markup=tagret_m)
    elif call.data == 'carta':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f'Выберите карту для оплаты', reply_markup=menu_card)
    elif call.data == 'oplata':
        menu_adm = types.InlineKeyboardMarkup(row_width=2)
        btn_sucsses = types.InlineKeyboardButton("Успешная оплата", callback_data=(f's_{user_id}'))
        btn_bad = types.InlineKeyboardButton("Оплата не прошла", callback_data=(f'b_{user_id}'))
        menu_adm.add(btn_bad, btn_sucsses)
        bot.send_message(admin_user, f'Проверить оплату от пользователя {user_id}', reply_markup=menu_adm)
        bot.send_message(user_id, "Менеджер обрабатывает ваш заказ")
    elif call.data.startswith('s_'):
        s = call.data[2:]
        bot.send_message(s, "Оплата успешна, спасибо!")
    elif call.data.startswith('b_'):
        s = call.data[2:]
        bot.send_message(s, "Оплата не прошла")
    elif call.data == 'back_dep':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Мой баланс💰💰💰", reply_markup=menu_balance)
    elif call.data == 'out':
        # Встановіть стан користувача
        user_state[user_id] = "waiting_for_card"
        bot.send_message(user_id, 'Введите номер карты или USDT(TRC20) Wallet')

        # Додайте обробник повідомлень
        @bot.message_handler(func=lambda message: True)
        def handle_message(message):
            user_id = message.chat.id
            if user_id in user_state:
                # Перевірка поточного стану користувача
                if user_state[user_id] == "waiting_for_card":
                    card = message.text
                    user_state[user_id] = "waiting_for_amount"
                    bot.send_message(user_id, 'Введите сумму для вывода')
                    user_data[user_id] = {"card": card}
                    return
                elif user_state[user_id] == "waiting_for_amount":
                    amount = message.text
                    user_state[user_id] = "none"
                    menu_out = types.InlineKeyboardMarkup(row_width=2)
                    btn_s = types.InlineKeyboardButton('Вывод прошёл', callback_data=f'o_{user_id}')
                    btn_b = types.InlineKeyboardButton('Вывод не прошёл', callback_data=f'ob_{user_id}')
                    menu_out.add(btn_b, btn_s)
                    try:
                        save = update_balance_xlsx(user_id, amount)
                        if save == False:
                            bot.send_message(user_id, "Сумма выводу превышает ваш баланс")
                        else:
                            bot.send_message(admin_user, f"Пользователь {user_id}\Реквизиты для выводу {user_data[user_id]['card']}\nСумма: {amount}", reply_markup=menu_out)
                            bot.send_message(user_id, "Менеджер обрабатывает Ваш запрос")
                    except:
                        bot.send_message(user_id, "При выводе роизошла ошибка")
                    return
            else:
                bot.send_message(user_id, "При обработке произошла ошибка")
                return
    elif call.data.startswith('o_'):
        s = call.data[2:]
        bot.send_message(s, "Вывод успешно")
    elif call.data.startswith('ob_'):
        s = call.data[3:]
        bot.send_message(s, "Вивід не пройшов")
    elif call.data == 'back':
        data = view_line()
        view_markup = types.InlineKeyboardMarkup(row_width=1)
        for key in data.keys():
                if len(data[key]) >= 3:  # Проверяем, что есть достаточно данных
                    name1, name2 = data[key][0], data[key][2]
                    if data[key][4]:
                        if key != list(data.keys())[-1] and data[key][4][0:4] == "яяяя":
                            view_markup.add(types.InlineKeyboardButton(f'{name1}   🆚   {name2}', callback_data=f"bet_{key}"))
        view_markup.add(types.InlineKeyboardButton("Назад", callback_data="back_s"))
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Ставки', reply_markup=view_markup)
    elif call.data == 'back_s':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Главное меню 🏠", reply_markup=menu_markup)
    elif call.data == 'history':
        try:
            wb = openpyxl.load_workbook(history)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)

            data_to_send = ""
            for row in rows:
                if row[0] == user_id:
                    data_to_send += f"|{row[0]}|{row[1]}|{row[2]}|\n"

            bot.send_message(user_id, data_to_send)
        except:
            bot.send_message(user_id, 'Нет записей')
    elif call.data == 'balance3':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Мой баланс💰💰💰", reply_markup=menu_balance)
    elif call.data == 'sber':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f'Номер карти: 2222222222, в реквизитах укажите {user_id}, \nНажмите на кнопку Оплатил и Ваш счёт будет пополнен на переведённую сумму в течении 1-15 минут', reply_markup=tagret_m)
    elif call.data == 'tink':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f'Номер карти: 2222222222, в реквизитах укажите {user_id}, \nНажмите на кнопку Оплатил и Ваш счёт будет пополнен на переведённую сумму в течении 1-15 минут', reply_markup=tagret_m)
    elif call.data == 'hist_out':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Выберите опцию", reply_markup=menu_out)
    elif call.data == 'dep_hist':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="История депозитов в разработке", reply_markup=menu_markup)
    elif call.data == 'out_hist':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="История выводов в разработке", reply_markup=menu_markup)
    elif call.data == 'back_to_cards':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Выберите карту для оплаты", reply_markup=menu_deposit)
    elif call.data == 'back_to_bal':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Мой баланс💰💰💰", reply_markup=menu_balance)

# Обробка Виклику меню
@bot.message_handler(func=lambda message: message.text == 'Вызов меню')
def menu(message):
    bot.send_message(message.chat.id, "Главное меню 🏠", reply_markup=menu_markup)

# Інструкція
@bot.message_handler(func=lambda message: message.text == 'Инструкция')
def instruction(message):
    instruction = """
Название бота: inkbet bot

Цель бота: возможность просто и быстро делать ставки без дополнительной верификации

Инструкция по использованию:

1. Запуск бота:
   - Запустите бота, нажав на кнопку "Start" или введя команду /start.

2. Регистрация/Логин:
   - Если вы новый пользователь, вам потребуется зарегистрироваться.
     Нажмите на кнопку регистрации и на Ваш баланс будет начислено 50000(в связи с розыгрышем)
   - Если Вы зарегистрировались раннее, нажмите на кнопку для логина

3. Баланс:
   - Для просмотра баланса: Мой баланс -> Посмотреть баланс

4. Пополнение счета:
   - Чтобы пополнить свой счет выберите кнопку депозита
   - Бот предоставит вам инструкции по пополнению счета через доступные методы

5. Сделать ставку:
   - Чтобы сделать ставку Вам необходимо выбрать Сделать ставку
   - Затем выберите одну из достурных ставок и введите сумму

7. Просмотр истории ставок:
   - История ставок ещё в разработке, но Вы можете посмотреть свои ставки в базовом формате

8. Вывод средств:
   - Для вывода средств выберите Вывод и следуйте инструкциям

9. Поддержка и помощь:
   - Если у вас возникли вопросы, проблемы или вам нужна дополнительная помощь выберите Обратная связь
"""
    bot.send_message(message.chat.id, instruction)

# Зворотній зв'язок   
@bot.message_handler(func=lambda message: message.text == "Обратная связь")
def call_message(message):
    bot.send_message(message.chat.id, "По всем вопросам обращаться @user")

# Информация о счете
@bot.message_handler(func=lambda message: message.text == 'Информация о счёте')
def menu(message):
    user_name = message.from_user.first_name
    user_nickname = message.from_user.username if message.from_user.username else "Нет никнейма"
    iser_id = message.from_user.id

    # Создание сообщения с User ID и никнеймом
    
    message_text = f"""Главное меню 🏠\n
🙋🏻‍♂️ Ник: {user_nickname}
👤 ID: {iser_id}
🎰 Активных ставок: в разработке"""
    # Отправка сообщения
    bot.send_message(message.chat.id, message_text, reply_markup=menu_markup)

import signal

# Обработчик сигнала прерывания
def signal_handler(signal, frame):
    print("\nПрограма завершена по сигналу прерывания (Ctrl+C).")
    exit(0)

# Зарегистрировать обработчик сигнала прерывания
signal.signal(signal.SIGINT, signal_handler)

while True:
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(f"Ошибка: {e}. Подключение через 3 секунды.")
        time.sleep(3)

